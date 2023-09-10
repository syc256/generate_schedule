import shutil
import sys
import os
import re
import xlrd
import xlwt
import logging
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QListWidget,
    QFileDialog,
    QMessageBox,
    QMenu,
    QAction,
    QLabel,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QLineEdit, QHBoxLayout,
)

# 定义常量
TOTAL_WEEKS = 16
COMPARE_DATE = '2023-09-01'
LOG_FILE = 'error_log.txt'


# 初始化日志记录器
def initialize_logger():
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    logging.basicConfig(
        filename=LOG_FILE,
        level=logging.ERROR,
        format='%(asctime)s [%(levelname)s]: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )


# 提取匹配的行
def find_pattern_lines(text, pattern):
    lines = text.split('\n')
    matching_lines = [line for line in lines if re.search(pattern, line)]
    return matching_lines


# 创建空白的无课表
def create_empty_schedule():
    empty_schedule = []
    for _ in range(1, TOTAL_WEEKS + 2):
        per_week_list = [
            [[], [], [], [], [], [], [], []],
            [[], [], [], [], [], [], [], []],
            [[], [], [], [], [], [], [], []],
            [[], [], [], [], [], [], [], []],
            [[], [], [], [], [], [], [], []],
            [[], [], [], [], [], [], [], []],
        ]
        empty_schedule.append(per_week_list)
    return empty_schedule


# 清空output文件夹
def clear_output_folder(output_folder):
    for file_name in os.listdir(output_folder):
        file_path = os.path.join(output_folder, file_name)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

# 设置界面
class SettingWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('设置')
        self.setGeometry(300, 300, 300, 200)
        self.week_line = QLineEdit()
        self.date_line = QLineEdit()
        self.initUI()

    def initUI(self):
        layout1 = QHBoxLayout()
        layout2 = QHBoxLayout()
        layout3 = QVBoxLayout()

        week_label = QLabel('周数设置：', self)

        self.week_line.setPlaceholderText("16")

        date_label = QLabel('最早打印时间：', self)

        self.date_line.setPlaceholderText('2023-09-01')

        self.confirm_button = QPushButton('确定')
        self.confirm_button.clicked.connect(self.get_input)

        layout1.addWidget(week_label)
        layout1.addWidget(self.week_line)
        layout2.addWidget(date_label)
        layout2.addWidget(self.date_line)
        layout3.addLayout(layout2)
        layout3.addLayout(layout1)
        layout3.addWidget(self.confirm_button)

        # 创建一个空的QWidget并将布局管理器设置为其布局
        container = QWidget()
        container.setLayout(layout3)

        # 设置QWidget为窗口的中央组件
        self.setCentralWidget(container)

    def get_input(self):
        global TOTAL_WEEKS
        global COMPARE_DATE
        TOTAL_WEEKS = int(self.week_line.text())
        COMPARE_DATE = self.date_line.text()
        self.close()


# 人员管理界面
class PersonManagementWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("人员管理")
        self.setGeometry(300, 300, 450, 500)
        self.imported_files = []
        self.info_path = ''
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        import_button = QPushButton("导入人员信息", self)
        import_button.setGeometry(430, 20, 150, 30)
        import_button.clicked.connect(self.import_files)

        self.file_label = QLabel('选择的文件将在这里显示', self)
        self.file_label.setGeometry(20, 20, 390, 30)

        self.table = QTableWidget()
        self.table.setGeometry(20, 50, 560, 400)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(['姓名', '性别', '校区'])

        self.confirm_button = QPushButton('确定')
        self.confirm_button.clicked.connect(self.export_to_excel)
        self.confirm_button.clicked.connect(self.window_close)

        self.save_button = QPushButton('保存')
        self.save_button.clicked.connect(self.export_to_excel)

        # 将组件添加到布局管理器中
        layout.addWidget(self.file_label)
        layout.addWidget(import_button)
        layout.addWidget(self.table)
        layout.addWidget(self.save_button)
        layout.addWidget(self.confirm_button)

        # 创建一个空的QWidget并将布局管理器设置为其布局
        container = QWidget()
        container.setLayout(layout)

        # 设置QWidget为窗口的中央组件
        self.setCentralWidget(container)

    def mem_info(self):
        workbook = xlrd.open_workbook(self.info_path)
        sheet = workbook.sheet_by_index(0)

        self.table.setRowCount(sheet.nrows - 1)  # 减1是因为第一行通常包含标题
        print(sheet.nrows - 1)

        for row in range(1, sheet.nrows):
            name = sheet.cell_value(row, 0)
            gender = sheet.cell_value(row, 1)
            campus = sheet.cell_value(row, 2)

            name_item = QTableWidgetItem(name)
            gender_item = QTableWidgetItem(gender)
            campus_item = QTableWidgetItem(campus)

            self.table.setItem(row - 1, 0, name_item)
            self.table.setItem(row - 1, 1, gender_item)
            self.table.setItem(row - 1, 2, campus_item)

    def import_files(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        file_filter = "Excel Files (*.xls)"
        self.info_path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", file_filter, options=options)

        if self.info_path:
            file_name = os.path.basename(self.info_path)
            self.file_label.setText(f'选择的文件：{file_name}')
            self.mem_info()

    def export_to_excel(self):
        if not os.path.exists('./temp'):
            os.makedirs('./temp')

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet('Sheet1')

        for row in range(self.table.rowCount()):
            name_item = self.table.item(row, 0)
            gender_item = self.table.item(row, 1)
            campus_item = self.table.item(row, 2)

            name = name_item.text()
            gender = gender_item.text()
            campus = campus_item.text()

            worksheet.write(row, 0, name)
            worksheet.write(row, 1, gender)
            worksheet.write(row, 2, campus)

        workbook.save(f'./temp/memberinfo.xls')
        print(f'数据已成功导出')

    def window_close(self):
        self.close()

# 主界面
class FileImporterGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.output_folder = "./output"
        self.imported_files = []
        self.initUI()

    def initUI(self):
        self.setWindowTitle("无课表生成内测-曲艺舞监专用")
        self.setGeometry(200, 200, 600, 670)

        self.file_list_widget = QListWidget(self)
        self.file_list_widget.setGeometry(20, 70, 560, 500)

        import_button = QPushButton("导入文件", self)
        import_button.setGeometry(20, 20, 100, 30)
        import_button.clicked.connect(self.import_files)

        setting_button = QPushButton('设置', self)
        setting_button.setGeometry(480, 20, 100, 30)
        setting_button.clicked.connect(self.open_setting_window)
        self.setting_window = None

        generate_button = QPushButton("生成无课表", self)
        generate_button.setGeometry(480, 600, 100, 30)
        generate_button.clicked.connect(self.generate_schedule)

        info_button = QPushButton("人员管理", self)
        info_button.setGeometry(20, 600, 100, 30)
        info_button.clicked.connect(self.open_person_management_window)
        self.person_management_window = None  # 添加一个属性用于存储人员管理窗口对象

        self.file_list_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.file_list_widget.customContextMenuRequested.connect(self.show_context_menu)

    # 打开设置窗口
    def open_setting_window(self):
        if not self.setting_window:
            self.setting_window = SettingWindow()
        self.setting_window.show()

    # 打开人员管理窗口
    def open_person_management_window(self):
        if not self.person_management_window:
            self.person_management_window = PersonManagementWindow()
        self.person_management_window.show()

    # 右键菜单
    def show_context_menu(self, pos):
        menu = QMenu()
        delete_action = QAction("删除文件", self)
        delete_action.triggered.connect(self.delete_selected_files)
        menu.addAction(delete_action)
        menu.exec_(self.file_list_widget.mapToGlobal(pos))

    # 右键删除文件
    def delete_selected_files(self):
        selected_items = self.file_list_widget.selectedItems()
        if not selected_items:
            return

        for item in selected_items:
            file_path = item.text()
            if file_path in self.imported_files:
                self.imported_files.remove(file_path)
                self.file_list_widget.takeItem(self.file_list_widget.row(item))

    # 添加文件
    def import_files(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        file_filter = "Excel Files (*.xls)"
        selected_files, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", file_filter, options=options)

        if selected_files:
            self.imported_files.extend(selected_files)
            self.file_list_widget.clear()
            self.file_list_widget.addItems(self.imported_files)

    # 生成课表
    def generate_schedule(self):
        try:
            print(TOTAL_WEEKS)
            print(COMPARE_DATE)
            if not os.path.exists(self.output_folder):
                os.makedirs(self.output_folder)

            spare = create_empty_schedule()

            for file_name in self.imported_files:
                workbook = xlrd.open_workbook(file_name)
                table = workbook.sheet_by_name('Sheet1')
                text = table.row_values(0)[0]
                matches = re.findall(r'[\u4e00-\u9fa5]+', text)  # 匹配汉字范围

                name = ''
                if len(matches) >= 3:
                    name = matches[1]    # 匹配人名
                    print(name, end=' ')
                else:
                    logging.error(f"{file_name}表头格式错误，请检查xls文件是否从教务下载")

                date = table.row_values(1)[0][-10:]
                print(date[-10:])
                if date < COMPARE_DATE:
                    logging.error(f"课表打印时间早于2023-09-01: {file_name}, 姓名: {name}")

                schedule_row = [3, 4, 5, 6, 7]
                row = -1

                for n in schedule_row:
                    row += 1
                    this_row = table.row_values(n)

                    for i in range(1, 8):
                        pattern = r'\(\[周\]\)'
                        matched_lines = find_pattern_lines(this_row[i], pattern)
                        week_list = []

                        for week_string_temp in matched_lines:
                            week_string = week_string_temp[:-5]
                            week_list_temp = week_string.split(',')

                            for num in range(0, len(week_list_temp)):
                                if '-' in week_list_temp[num]:
                                    nums = week_list_temp[num].split('-')
                                    left = int(nums[0])
                                    right = int(nums[1])
                                    new_weeks = list(range(left, right + 1))
                                    week_list += new_weeks
                                else:
                                    week_list.append(int(week_list_temp[num]))
                        week_list.sort()
                        int_numbers = [int(x) for x in week_list]

                        remaining_list = [x for x in range(1, TOTAL_WEEKS + 1) if x not in int_numbers]

                        for week in remaining_list:
                            spare[week][row][i - 1].append(name + ' ')

            for i in range(1, TOTAL_WEEKS + 1):
                print(i)
                for j in range(6):
                    print(spare[i][j])

            for i in range(1, TOTAL_WEEKS + 1):
                workbook = xlwt.Workbook()
                worksheet = workbook.add_sheet('Sheet1')
                worksheet.col(0).width = 3000
                for j in range(1, 8):
                    worksheet.col(j).width = 7000

                borders = xlwt.Borders()
                borders.left = xlwt.Borders.THIN
                borders.right = xlwt.Borders.THIN
                borders.top = xlwt.Borders.THIN
                borders.bottom = xlwt.Borders.THIN
                style = xlwt.XFStyle()
                style.borders = borders
                style.alignment.wrap = 1

                weekday = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
                session = ['第一大节', '第二大节', '第三大节', '第四大节', '晚课']

                s = 0
                for j in weekday:
                    s += 1
                    worksheet.write(0, s, j, style)

                s = 0
                for j in session:
                    s += 1
                    worksheet.write(s, 0, j, style)

                for x in range(0, 5):
                    for y in range(0, 7):
                        worksheet.write(x + 1, y + 1, spare[i][x][y], style)
                workbook_path = f'{self.output_folder}/第' + str(i) + '周.xls'
                workbook.save(workbook_path)

            QMessageBox.information(self, "成功", f"无课表已生成，文件路径: {self.output_folder}")
        except Exception as e:
            QMessageBox.critical(self, "错误", str(e))


def main():
    initialize_logger()
    app = QApplication(sys.argv)
    window = FileImporterGUI()
    window.show()
    clear_output_folder(window.output_folder)  # 在创建window之后清空output文件夹
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
